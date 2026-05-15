"""
Parse messy Excel files and extract logical table structure.

Reads .xlsx files with openpyxl, classifies rows, detects table islands
via BFS flood-fill, and outputs structured JSON to stdout.
"""

import subprocess
import sys

for pkg in ["openpyxl"]:
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

import argparse
import json
import os
import re
from collections import deque
from statistics import median

import openpyxl
from openpyxl.utils import get_column_letter

MAX_ROWS = 10000

SUBTOTAL_RE = re.compile(
    r"合\s*计|小\s*计|总\s*计|汇\s*总|total|subtotal|sum|grand\s+total",
    re.IGNORECASE,
)
NOTES_RE = re.compile(
    r"^(备注|注|说明|note|remark|\*)", re.IGNORECASE
)


# ---------------------------------------------------------------------------
# Step 1: CellSignature extraction
# ---------------------------------------------------------------------------

def _resolve_bg_color(cell):
    """Return hex RGB string for background color, defaulting to FFFFFF."""
    try:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.rgb:
            rgb = str(fill.fgColor.rgb)
            if rgb in ("00000000", "0", "00000000"):
                return "FFFFFF"
            # Strip alpha prefix if present (e.g. "FF4472C4" -> "4472C4")
            if len(rgb) == 8:
                return rgb[2:]
            return rgb
    except Exception:
        pass
    return "FFFFFF"


def _resolve_alignment(cell):
    """Return alignment as one of center|left|right|general."""
    try:
        if cell.alignment and cell.alignment.horizontal:
            h = cell.alignment.horizontal.lower()
            if h in ("center", "left", "right"):
                return h
    except Exception:
        pass
    return "general"


def cell_signature(cell, merged_map):
    """Build a CellSignature dict for a single cell."""
    value = cell.value
    is_empty = value is None or (isinstance(value, str) and value.strip() == "")
    is_merged = (cell.row, cell.column) in merged_map
    merge_span = merged_map.get((cell.row, cell.column), [1, 1])

    is_bold = False
    font_size = 11.0
    try:
        if cell.font:
            is_bold = bool(cell.font.bold)
            if cell.font.size is not None:
                font_size = float(cell.font.size)
    except Exception:
        pass

    bg_color = _resolve_bg_color(cell)
    alignment = _resolve_alignment(cell)

    number_format = ""
    try:
        if cell.number_format and cell.number_format != "General":
            number_format = str(cell.number_format)
    except Exception:
        pass

    has_formula = False
    if isinstance(value, str) and value.startswith("="):
        has_formula = True

    return {
        "value": value,
        "is_empty": is_empty,
        "is_merged": is_merged,
        "merge_span": merge_span,
        "is_bold": is_bold,
        "font_size": font_size,
        "bg_color": bg_color,
        "alignment": alignment,
        "number_format": number_format,
        "has_formula": has_formula,
    }


# ---------------------------------------------------------------------------
# Step 2: Expand merged cells & build merged_map
# ---------------------------------------------------------------------------

def build_merged_map(ws):
    """Return dict mapping (row, col) -> [row_span, col_span] for every cell
    in a merged range.  Also return a set of top-left cells."""
    merged_map = {}
    top_lefts = set()
    for rng in ws.merged_cells.ranges:
        min_r, min_c = rng.min_row, rng.min_col
        max_r, max_c = rng.max_row, rng.max_col
        row_span = max_r - min_r + 1
        col_span = max_c - min_c + 1
        top_lefts.add((min_r, min_c))
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) == (min_r, min_c):
                    merged_map[(r, c)] = [row_span, col_span]
                else:
                    merged_map[(r, c)] = [1, 1]
    return merged_map, top_lefts


def expand_merged_cells(ws, grid, merged_map):
    """Copy the top-left value of every merged range into all cells of
    the range inside *grid*.  Mark expanded cells."""
    for rng in ws.merged_cells.ranges:
        min_r, min_c = rng.min_row, rng.min_col
        max_r, max_c = rng.max_row, rng.max_col
        tl_sig = grid.get((min_r, min_c))
        if tl_sig is None:
            continue
        val = tl_sig["value"]
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    if (r, c) in grid:
                        grid[(r, c)]["value"] = val
                        grid[(r, c)]["is_empty"] = val is None or (
                            isinstance(val, str) and val.strip() == ""
                        )
                        grid[(r, c)]["_is_expanded"] = True


# ---------------------------------------------------------------------------
# Step 3: Row type classification
# ---------------------------------------------------------------------------

def _is_numeric_like(val):
    """Return True if value looks numeric (int, float, or numeric string)."""
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        s = val.strip().replace(",", "").replace("%", "")
        try:
            float(s)
            return True
        except (ValueError, TypeError):
            return False
    return False


def classify_rows(grid, total_rows, total_cols, median_fs):
    """Classify each row (1-indexed) and return dict row_num -> type."""
    classifications = {}

    for row in range(1, total_rows + 1):
        sigs = [grid.get((row, c)) for c in range(1, total_cols + 1)]
        non_empty = [s for s in sigs if s and not s["is_empty"]]

        # EMPTY
        if not non_empty:
            classifications[row] = "EMPTY"
            continue

        # SUBTOTAL — keyword must appear in the first non-empty cell (row label)
        # or the row must have ≤ 2 non-formula string cells containing the keyword.
        # This avoids marking header rows where "合计" is just a column name.
        first_ne_val = str(non_empty[0]["value"]) if non_empty[0]["value"] is not None else ""
        if SUBTOTAL_RE.search(first_ne_val) and not non_empty[0]["has_formula"]:
            classifications[row] = "SUBTOTAL"
            continue

        # NOTES
        if len(non_empty) <= 2:
            first_val = str(non_empty[0]["value"]).strip() if non_empty[0]["value"] is not None else ""
            if NOTES_RE.match(first_val):
                classifications[row] = "NOTES"
                continue

        # TITLE
        has_wide_merge = any(
            s["is_merged"] and s["merge_span"][1] >= 3 for s in non_empty
        )
        has_large_font = median_fs > 0 and any(
            s["font_size"] > median_fs * 1.15 and s["alignment"] == "center"
            for s in non_empty
        )
        all_non_numeric = all(not _is_numeric_like(s["value"]) for s in non_empty)
        if len(non_empty) <= 3 and (has_wide_merge or has_large_font or (all_non_numeric and row <= 5)):
            classifications[row] = "TITLE"
            continue

        # Defer HEADER vs DATA — need lookahead
        classifications[row] = None  # placeholder

    # Second pass: HEADER vs DATA with lookahead
    rows_ordered = sorted(classifications.keys())
    for idx, row in enumerate(rows_ordered):
        if classifications[row] is not None:
            continue

        sigs = [grid.get((row, c)) for c in range(1, total_cols + 1)]
        non_empty = [s for s in sigs if s and not s["is_empty"]]
        string_count = sum(1 for s in non_empty if not _is_numeric_like(s["value"]))
        string_ratio = string_count / len(non_empty) if non_empty else 0

        # Find the next non-EMPTY row (skip rows already classified as EMPTY;
        # accept both resolved non-EMPTY rows and still-unclassified None rows
        # so that multi-row header lookahead works in small tables)
        next_row = None
        for nr in rows_ordered[idx + 1:]:
            cls = classifications.get(nr)
            if cls != "EMPTY":
                next_row = nr
                break

        next_is_numeric = False
        if next_row is not None:
            next_sigs = [grid.get((next_row, c)) for c in range(1, total_cols + 1)]
            next_non_empty = [s for s in next_sigs if s and not s["is_empty"]]
            if next_non_empty:
                numeric_count = sum(1 for s in next_non_empty if _is_numeric_like(s["value"]))
                next_is_numeric = numeric_count / len(next_non_empty) > 0.5

        if (
            string_ratio >= 0.5
            and next_is_numeric
            and len(non_empty) >= 2
        ):
            classifications[row] = "HEADER"
        else:
            classifications[row] = "DATA"

    return classifications


# ---------------------------------------------------------------------------
# Step 4: 2D island detection via BFS flood-fill
# ---------------------------------------------------------------------------

def detect_islands(grid, total_rows, total_cols):
    """BFS flood-fill on occupied cells (4-connectivity).
    Returns list of (min_row, max_row, min_col, max_col) 1-indexed."""
    occupied = set()
    for (r, c), sig in grid.items():
        if not sig["is_empty"]:
            occupied.add((r, c))

    visited = set()
    islands = []

    for cell in occupied:
        if cell in visited:
            continue
        # BFS
        queue = deque([cell])
        visited.add(cell)
        region_cells = [cell]
        while queue:
            cr, cc = queue.popleft()
            for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                nr, nc = cr + dr, cc + dc
                if (nr, nc) in occupied and (nr, nc) not in visited:
                    visited.add((nr, nc))
                    queue.append((nr, nc))
                    region_cells.append((nr, nc))

        min_r = min(r for r, c in region_cells)
        max_r = max(r for r, c in region_cells)
        min_c = min(c for r, c in region_cells)
        max_c = max(c for r, c in region_cells)

        row_span = max_r - min_r + 1
        col_span = max_c - min_c + 1
        if row_span >= 3 and col_span >= 2:
            islands.append((min_r, max_r, min_c, max_c))

    # Sort by position (top-left first)
    islands.sort(key=lambda x: (x[0], x[2]))
    return islands


# ---------------------------------------------------------------------------
# Step 5: Extract logical tables from each island
# ---------------------------------------------------------------------------

def extract_tables(islands, row_classifications, grid, total_cols):
    """For each island, extract a logical table definition."""
    tables = []

    for table_idx, (min_r, max_r, min_c, max_c) in enumerate(islands):
        rows_in_region = list(range(min_r, max_r + 1))
        header_rows = [r for r in rows_in_region if row_classifications.get(r) == "HEADER"]
        data_rows = [r for r in rows_in_region if row_classifications.get(r) == "DATA"]
        title_rows = [r for r in rows_in_region if row_classifications.get(r) == "TITLE"]
        empty_rows = [r for r in rows_in_region if row_classifications.get(r) == "EMPTY"]
        subtotal_rows = [r for r in rows_in_region if row_classifications.get(r) == "SUBTOTAL"]
        notes_rows = [r for r in rows_in_region if row_classifications.get(r) == "NOTES"]

        # Primary header: last HEADER row before first DATA row
        primary_header = None
        first_data = data_rows[0] if data_rows else None
        if header_rows and first_data is not None:
            candidates = [h for h in header_rows if h < first_data]
            if candidates:
                primary_header = candidates[-1]
        elif header_rows:
            primary_header = header_rows[-1]

        # Multi-row header unification
        multi_header_rows = []
        if primary_header is not None and first_data is not None:
            multi_header_rows = [h for h in header_rows if h <= primary_header]
        elif header_rows:
            multi_header_rows = header_rows

        # Build columns
        columns = _build_columns(multi_header_rows, min_c, max_c, grid)

        # Ambiguous rows: could be HEADER or DATA (mixed content)
        ambiguous = []
        for r in rows_in_region:
            cls = row_classifications.get(r)
            if cls == "DATA":
                sigs = [grid.get((r, c)) for c in range(min_c, max_c + 1)]
                non_empty = [s for s in sigs if s and not s["is_empty"]]
                if non_empty:
                    string_count = sum(1 for s in non_empty if not _is_numeric_like(s["value"]))
                    ratio = string_count / len(non_empty)
                    if 0.3 <= ratio <= 0.7:
                        ambiguous.append(r)

        # Confidence calculation
        confidence = _calc_confidence(
            header_rows, data_rows, ambiguous, rows_in_region,
            row_classifications, min_c, max_c, grid,
        )

        # Row classifications subset
        row_cls_map = {}
        for r in rows_in_region:
            row_cls_map[str(r)] = row_classifications.get(r, "DATA")

        excluded = {
            "title": title_rows,
            "meta": [],
            "empty": empty_rows,
            "subtotal": subtotal_rows,
            "notes": notes_rows,
        }

        tables.append({
            "table_index": table_idx,
            "region": {
                "min_row": min_r,
                "max_row": max_r,
                "min_col": min_c,
                "max_col": max_c,
            },
            "header_rows": header_rows,
            "data_rows": data_rows,
            "excluded_rows": excluded,
            "columns": columns,
            "confidence": round(confidence, 2),
            "ambiguous_rows": ambiguous,
            "row_classifications": row_cls_map,
        })

    return tables


def _build_columns(header_rows, min_c, max_c, grid):
    """Build column definitions from (possibly multi-row) headers."""
    columns = []
    for c in range(min_c, max_c + 1):
        parts = []
        for r in header_rows:
            sig = grid.get((r, c))
            if sig and not sig["is_empty"] and sig["value"] is not None:
                parts.append(str(sig["value"]).strip())
        name = "_".join(parts) if parts else f"Col_{c}"
        columns.append({
            "name": name,
            "col_index": c,
            "inferred_type": "unknown",
        })
    return columns


def _calc_confidence(
    header_rows, data_rows, ambiguous, rows_in_region,
    row_classifications, min_c, max_c, grid,
):
    """Calculate confidence score for a logical table."""
    conf = 0.5

    if len(header_rows) == 1:
        conf += 0.2

    if len(data_rows) > 3:
        conf += 0.1

    if not ambiguous:
        conf += 0.1

    # Check header count consistency with data rows
    if header_rows:
        h = header_rows[-1]
        h_sigs = [grid.get((h, c)) for c in range(min_c, max_c + 1)]
        h_nonempty = sum(1 for s in h_sigs if s and not s["is_empty"])
        if data_rows:
            d = data_rows[0]
            d_sigs = [grid.get((d, c)) for c in range(min_c, max_c + 1)]
            d_nonempty = sum(1 for s in d_sigs if s and not s["is_empty"])
            if h_nonempty > 0 and abs(h_nonempty - d_nonempty) <= 1:
                conf += 0.1

    # Disconnected header candidates penalty
    if len(header_rows) >= 2:
        gaps = 0
        for i in range(1, len(header_rows)):
            if header_rows[i] - header_rows[i - 1] > 1:
                # Check if rows between are non-header
                between = range(header_rows[i - 1] + 1, header_rows[i])
                if any(row_classifications.get(r) not in ("HEADER", "EMPTY") for r in between):
                    gaps += 1
        if gaps > 0:
            conf -= 0.2

    # Mixed row types penalty
    types_in_region = set(
        row_classifications.get(r) for r in rows_in_region
        if row_classifications.get(r) not in ("EMPTY",)
    )
    if len(types_in_region) >= 4:
        conf -= 0.1

    return max(0.0, min(1.0, conf))


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def process_sheet(ws):
    """Process a single worksheet and return its structure dict."""
    total_rows = min(ws.max_row or 0, MAX_ROWS)
    total_cols = ws.max_column or 0

    if total_rows == 0 or total_cols == 0:
        return {
            "name": ws.title,
            "total_rows": total_rows,
            "total_cols": total_cols,
            "logical_tables": [],
        }

    # Build merged map
    merged_map, top_lefts = build_merged_map(ws)

    # Extract cell signatures into grid
    grid = {}
    for row in ws.iter_rows(min_row=1, max_row=total_rows, min_col=1, max_col=total_cols):
        for cell in row:
            grid[(cell.row, cell.column)] = cell_signature(cell, merged_map)

    # Expand merged cells
    expand_merged_cells(ws, grid, merged_map)

    # Compute median font size across non-empty cells
    font_sizes = [
        sig["font_size"]
        for sig in grid.values()
        if not sig["is_empty"]
    ]
    median_fs = median(font_sizes) if font_sizes else 11.0

    # Classify rows
    row_classifications = classify_rows(grid, total_rows, total_cols, median_fs)

    # Detect islands
    islands = detect_islands(grid, total_rows, total_cols)

    # Extract logical tables
    logical_tables = extract_tables(islands, row_classifications, grid, total_cols)

    return {
        "name": ws.title,
        "total_rows": total_rows,
        "total_cols": total_cols,
        "logical_tables": logical_tables,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Parse Excel file structure and extract logical tables."
    )
    parser.add_argument("--file", required=True, help="Path to .xlsx file")
    parser.add_argument("--sheet", default=None, help="Sheet name (omit to process all)")
    args = parser.parse_args()

    file_path = args.file
    if not os.path.isfile(file_path):
        result = {
            "file": os.path.basename(file_path),
            "error": f"File not found: {file_path}",
            "sheets": [],
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False)
    except Exception as e:
        result = {
            "file": os.path.basename(file_path),
            "error": str(e),
            "sheets": [],
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(1)

    sheets_to_process = []
    if args.sheet:
        if args.sheet in wb.sheetnames:
            sheets_to_process = [wb[args.sheet]]
        else:
            result = {
                "file": os.path.basename(file_path),
                "error": f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}",
                "sheets": [],
            }
            print(json.dumps(result, ensure_ascii=False, indent=2))
            sys.exit(1)
    else:
        sheets_to_process = [wb[name] for name in wb.sheetnames]

    sheet_results = []
    for ws in sheets_to_process:
        try:
            sheet_results.append(process_sheet(ws))
        except Exception as e:
            print(f"Error processing sheet '{ws.title}': {e}", file=sys.stderr)
            sheet_results.append({
                "name": ws.title,
                "total_rows": 0,
                "total_cols": 0,
                "logical_tables": [],
            })

    output = {
        "file": os.path.basename(file_path),
        "sheets": sheet_results,
    }
    print(json.dumps(output, ensure_ascii=False, indent=2))
    sys.exit(0)


if __name__ == "__main__":
    main()
